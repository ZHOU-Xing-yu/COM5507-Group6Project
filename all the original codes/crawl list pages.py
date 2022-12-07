

# 提取方法1：指针头（page=【1：13】） 列表 for循环 输出pandas

# 提取方法2：selenium自动翻页循环解析
# 1. 解析页面元素
# 2. 列表页内容提取 （正则表达式 + XPath + bs选择器）
#    使用xpath（以上都行）封装html源码解析即可
# 3. 点击'下一页'循环以上操作
# 4. 追加保存数据到Excel表
from selenium.webdriver import ChromeOptions
from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
from os import makedirs
from os.path import exists
import bs4 as bs
import pandas as pd
from openpyxl import load_workbook
import time
import re
from lxml import etree
import random

options = ChromeOptions()
options.add_experimental_option('excludeSwitches', ['enable-automation'])
options.add_experimental_option('useAutomationExtension', False)
        # 此步骤很重要，设置为开发者模式，防止被各大网站识别出来使用了Selenium

browser = webdriver.Chrome(executable_path="/Users/zhouxingyu/Desktop/PYTHONSTUDY/chromedriver 2", options=options)
browser.execute_cdp_cmd('Page.addScriptToEvaluateOnNewDocument', {
     'source': 'Object.defineProperty(navigator, "webdriver", {get: () => undefined})'
})

INDEX_URL = 'https://web3.career/web3-non-tech-salaries'
TOTAL_PAGE = 14     # 共爬取14页
RESULTS_DIR = 'results'  # 结果文件夹
RESULTS_FILE = 'web3job(non-tech).xlsx'

exists(RESULTS_DIR) or makedirs(RESULTS_DIR)  # 文件夹不存在就新建

# 选择方法：CSS选择器 xpath（可直接复制黏贴） ID选择器 CLASS_NAME

def crawlIndex():
    html=""
    try:
        browser.get(INDEX_URL)
        html = browser.page_source
    except TimeoutException:
        print("访问超时")
    # with open(f"{RESULTS_DIR}/web3career.html", 'w', encoding="utf8") as f:
    #     f.write(html)
    return html

def parseList(html):
    #doc = etree.HTML(html)
    #res = doc.xpath('//div/div/div/a/h3')
    html = browser.page_source
    page_html = bs.BeautifulSoup(html, 'html.parser')
    names = page_html.find_all('h2', class_='fs-6 fs-md-5 fw-bold my-primary')
    companies = page_html.find_all(class_="mt-auto d-block d-md-flex")
    locations = page_html.find_all('td', class_='job-location-mobile')
    salaries = page_html.find_all('p', class_='ps-0 mb-0 text-salary')
    classifications = page_html.find_all('td', class_='align-middle d-none d-md-table-cell')
    company = []
    name = []
    location = []
    salary = []
    classification = []
    for i in companies:
        company.append(i.get_text(strip=True))
    for i in names:
        name.append(i.get_text(strip=True))
    for i in locations:
        location.append(i.get_text(strip=True))
    for i in salaries:
        salary.append(i.get_text(strip=True))
    for i in classifications:
        classification.append(i.get_text(strip=True))
        for item in classification:
            if "Apply" in item:
                classification.remove(item)
    return list(zip(name,company,location,salary,classification))



def crawlList():
    html = ""
    try:
        nextPage = WebDriverWait(browser, 5).until(EC.element_to_be_clickable((By.XPATH, "//li[@class='page-item next']")))
        nextPage.click()
        WebDriverWait(browser, 5).until(EC.presence_of_element_located((By.CLASS_NAME, 'my-btn my-btn-primary-reserve')))
        html = browser.page_source
        # print(html)
    except TimeoutException:
        print("访问超时2")
    return html



def saveData(data):
    header = ['NAME', 'COMPANY','LOCATION', 'SALARY','CLASSIFICATION']
    #, 'NAME', 'LOCATION', 'SALARY'
    # create a Pandas frame from a dict
    # guideline: https://pandas.pydata.org/pandas-docs/stable/generated/pandas.DataFrame.from_dict.html
    df = pd.DataFrame(data, columns=header)
    if exists(f'{RESULTS_DIR}/{RESULTS_FILE}'):
        wb = load_workbook(filename=f'{RESULTS_DIR}/{RESULTS_FILE}')  # 实例化excel
        sh = wb.active  # 获取sheet
        all_excel_data = list(sh.iter_rows(values_only=True, min_row=2))  # 获取表格所有数据,忽略表头
        # print(all_excel_data,type(all_excel_data))
        original_data = pd.DataFrame(all_excel_data, columns=header)
        save_data = pd.concat([original_data, df])  # 合并两个数据表

        save_data.to_excel(f'{RESULTS_DIR}/{RESULTS_FILE}', index=False)
    else:
        df.to_excel(f'{RESULTS_DIR}/{RESULTS_FILE}', index=False, header=header)


def main():
    html = crawlIndex()
    #print(html)
    # with open(f'{RESULTS_DIR}/web3career.html', 'r', encoding='utf8') as f:
    #     html = f.read()
    for i in range(TOTAL_PAGE):
        jobsData = parseList(html)
        saveData(jobsData)
        html = crawlList()  # 点击“下一页”按钮，获取下一页源码
    browser.quit()


if __name__ == '__main__':
    main()